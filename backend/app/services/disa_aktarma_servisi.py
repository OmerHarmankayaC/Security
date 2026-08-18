"""Excel disa aktarma (SDD 5.8; SRS FR-8.5, FR-8.9).

VERI MEVCUT OKUMA YUZEYLERINDEN GELIR, IKINCI BIR HESAP YAPILMAZ. Cizelge
atamalardan, analiz `AnalizServisi`'nden, aciklar kapsama kayitlarindan,
fazla calisma H10'un kendi fonksiyonundan okunur. Disa aktarmanin kendi
toplamlarini hesaplamasi, ayni sayinin ekranda ve dosyada farkli cikmasi
demektir; bu projede ayni hesabin iki yerde durmasinin bedeli birkac kez
odenmistir.

SAAT BICIMLEMESI DE TEK YERDEN: `zaman_araligi.saat_metni` ve
`aralik_metni`. Saat metni biciminin ucuncu bir kopyasi bir kez hataya yol
acti.

HUCRE DOLGUSU BILGIYI TEK BASINA TASIMAZ. Saat araligi hucrede METIN olarak
da yazilidir ve bir aciklama satiri dolgunun anlamini soyler; renksiz
basilan bir cikti okunabilir kalir. Ayni ilke ekrandaki renk bandi icin de
gecerlidir (SDD 6.3.3).
"""

from dataclasses import dataclass
from datetime import date, datetime, timedelta

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from sqlalchemy.orm import Session

from app.kurallar.kayit_defteri import kurallari_yukle
from app.kurallar.zaman_araligi import aralik_metni, gece_saat_sayisi, saat_metni
from app.kurallar.zorunlu import h10_fazla_calisma_saatleri
from app.models.sonuc import CizelgeSurumu
from app.repositories.kural import KuralDeposu
from app.repositories.sonuc import (
    AtamaDeposu,
    CizelgeSurumuDeposu,
    DonemDeposu,
    FazlaKadroDeposu,
    KapsamaAcigiDeposu,
)
from app.repositories.tanim import GorevNoktasiDeposu, PersonelDeposu
from app.schemas.analiz import AnalizOku
from app.services.analiz_servisi import AnalizServisi
from app.services.atama_donusumu import atama_kayitlarina_cevir
from app.services.baglam_kurucu import baglam_olustur

# --- Bicimleme sabitleri --------------------------------------------------

# Renkler 8 HANELI (alfa + RGB) yazilir. openpyxl 6 haneli bir deger
# aldiginda alfayi 00 kabul ediyor ve Excel dolguyu saydam gosteriyor.
_MARKA = "FF1F4E45"  # baslik bandi ve sayfa basligi
_SESSIZ = "FF6B6B60"  # ikincil metin
_VURGU = "FFC1502E"  # karsilanmayan talep gibi DIKKAT isteyen sayilar
_BEYAZ = "FFFFFFFF"
_KOYU_METIN = "FF1A1A16"
_BANT = "FFF7F6F0"  # satir bantlamasinin koyu adimi

# SAAT BANDI: dolgu blogun BASLADIGI saatten gelir; 13.00 en acik, 01.00 en
# koyu ve arasi surekli gecer. Ekrandaki renk bandinin (SDD 6.3.3) hucreye
# indirgenmis halidir - orada bir gradient, burada 13 adim. Onceki surum
# "gece / kismen gece / gunduz" diye UC kovaya bolen AYRI bir olcu
# kullaniyordu; ayni bilgi iki yerde iki farkli bicimde tanimlanmis oluyordu.
# Liste 13.00'ten uzaklik (0-12) ile indislenir.
_SAAT_BANDI = (
    "FFE9E7D9",
    "FFE5E4D6",
    "FFDCDBCE",
    "FFCDCDC1",
    "FFBABBB0",
    "FFA4A69D",
    "FF8C9088",
    "FF737A73",
    "FF5D6560",
    "FF4A534F",
    "FF3B4542",
    "FF323C3A",
    "FF2F3A38",
)
_EN_ACIK_SAAT = 13
_KOYU_ESIK = 7  # bu adimdan itibaren dolgu koyu, yazi beyaza doner

_ACIK_DOLGU = PatternFill("solid", fgColor="FFF7E2D6")  # kapsama acigi olan gun
_BASLIK_DOLGU = PatternFill("solid", fgColor=_MARKA)
_BANT_DOLGU = PatternFill("solid", fgColor=_BANT)
_DUZ_DOLGU = PatternFill("solid", fgColor=_BEYAZ)

_BASLIK_YAZI = Font(bold=True, size=10, color=_BEYAZ)
_SAYFA_BASLIGI = Font(bold=True, size=16, color=_MARKA)
_ALT_BASLIK = Font(size=9, color=_SESSIZ)
_OLCU_YAZI = Font(bold=True, size=10, color=_VURGU)
_ETIKET_YAZI = Font(size=10, color=_SESSIZ)
_NOT_YAZI = Font(size=8, color=_SESSIZ)
_BUYUK_SAYI = Font(bold=True, size=20, color=_MARKA)
_BUYUK_VURGU = Font(bold=True, size=20, color=_VURGU)
_VERI_YAZI = Font(size=10)
_KOYU_VERI = Font(bold=True, size=10)
_HAM_YAZI = Font(size=9)

_INCE = Side(style="thin", color="FFD2D7CE")
_KENARLIK = Border(left=_INCE, right=_INCE, top=_INCE, bottom=_INCE)
_ORTALI = Alignment(horizontal="center", vertical="center", wrap_text=True)
_SOLA = Alignment(vertical="center")


def _tarih_tr(gun: date) -> str:
    """`20.07.2026` - dosyanin okuyucusu INSANDIR.

    Ham veri sayfasi ISO kalir (SRS 7.2): orasi makine icindir ve ISO
    sozluksel siralamada takvim sirasiyla aynidir. Ikisinin ayni dosyada
    farkli bicimde durmasi bilinclidir.
    """
    return gun.strftime("%d.%m.%Y")


def _yuzde_tr(oran: float | None) -> str:
    """`%96,9` - ondalik ayraci VIRGUL (Turkce yerel)."""
    return "—" if oran is None else f"%{oran * 100:.1f}".replace(".", ",")


def _nokta_kisaltmasi(ad: str) -> str:
    """`Guvenlik` -> `GÜV`. Tam ad hucreye sigmaz; kisaltma ayirt edicidir."""
    return ad[:3].upper() if ad else ""


def _kisi_saat(baslangic: datetime, bitis: datetime, sayi: int) -> int:
    """Sapmanin KISI-SAAT olcusu: eksik kisi x aralik uzunlugu.

    Aralik sayisi ile kisi-saat FARKLI olculerdir ve ikisi de raporda durur:
    ardisik saatler tek kayitta birlestigi icin (SDD 4.2.4) "10 aralik"
    kaydin sayisini, "36 kisi-saat" gercek eksigin buyuklugunu soyler.
    """
    return sayi * round((bitis - baslangic).total_seconds() / 3600)


def _blok_dolgusu(baslangic: datetime) -> tuple[PatternFill, Font]:
    """Blogun BASLADIGI saate gore dolgu ve okunur yazi rengi.

    Renk BILGI TASIMAZ - saat araligi hucrede zaten yazili - yalnizca goz
    gezdirmeyi kolaylastirir. Koyu adimlarda yazi beyaza doner; yoksa metin
    kaybolur ve "renksiz de okunur" sozu bozulur.
    """
    saat = baslangic.hour
    uzaklik = min((saat - _EN_ACIK_SAAT) % 24, (_EN_ACIK_SAAT - saat) % 24)
    dolgu = PatternFill("solid", fgColor=_SAAT_BANDI[uzaklik])
    renk = _BEYAZ if uzaklik >= _KOYU_ESIK else _KOYU_METIN
    return dolgu, Font(bold=True, size=9, color=renk)


def _baslik_satiri(sayfa: Worksheet, satir: int, basliklar: list[str]) -> None:
    """Koyu yesil sutun basligi bandi - HER SAYFADA AYNI."""
    for sutun, metin in enumerate(basliklar, start=1):
        hucre = sayfa.cell(satir, sutun, metin)
        hucre.font = _BASLIK_YAZI
        hucre.fill = _BASLIK_DOLGU
        hucre.alignment = _ORTALI
    sayfa.row_dimensions[satir].height = 30


def _bantla(sayfa: Worksheet, satir: int, sutun_sayisi: int, tek: bool) -> None:
    """Satir bantlamasi: uzun listede goz satiri kaybetmesin."""
    dolgu = _BANT_DOLGU if tek else _DUZ_DOLGU
    for sutun in range(1, sutun_sayisi + 1):
        hucre = sayfa.cell(satir, sutun)
        hucre.fill = dolgu
        hucre.font = _VERI_YAZI
    sayfa.row_dimensions[satir].height = 15


@dataclass(frozen=True, slots=True)
class _Baglam:
    """Iki calisma kitabinin da okudugu ortak veri."""

    surum: CizelgeSurumu
    donem_baslangic: date
    donem_bitis: date
    gunler: list[date]
    atamalar: list
    aciklar: list
    fazlalar: list
    analiz: AnalizOku
    personel_adi: dict[int, str]
    personel_sicil: dict[int, str]
    nokta_adi: dict[int, str]
    fazla_calisma: dict[int, float]
    kalan_kota: dict[int, float]
    # ARALIK SAYISI ile KISI-SAAT ayri olculerdir; ikisi de raporda durur.
    karsilanmayan_kisi_saat: int
    ceza_kalemleri: list[tuple[str, str, float, float]]


class DisaAktarmaServisi:
    def __init__(self, oturum: Session) -> None:
        self.oturum = oturum
        self.surum = CizelgeSurumuDeposu(oturum)
        self.donem = DonemDeposu(oturum)
        self.atama = AtamaDeposu(oturum)
        self.kapsama = KapsamaAcigiDeposu(oturum)
        self.fazla = FazlaKadroDeposu(oturum)
        self.personel = PersonelDeposu(oturum)
        self.nokta = GorevNoktasiDeposu(oturum)
        self.kural = KuralDeposu(oturum)
        self.analiz = AnalizServisi(oturum)

    def _baglami_kur(self, surum_id: int) -> _Baglam | None:
        surum = self.surum.getir(surum_id)
        if surum is None:
            return None
        donem = self.donem.getir(surum.donem_id)
        if donem is None:
            return None
        analiz = self.analiz.hesapla(surum_id)
        if analiz is None:
            return None

        atamalar = list(self.atama.surume_gore_getir(surum_id))
        gun_sayisi = (donem.bitis_tarihi - donem.baslangic_tarihi).days + 1
        gunler = [donem.baslangic_tarihi + timedelta(days=i) for i in range(gun_sayisi)]

        # FAZLA CALISMA VE KOTA H10'UN KENDI FONKSIYONUNDAN. Burada yeniden
        # hesaplansaydi dosyada, sistemin baska hicbir yerinde bulunmayan bir
        # sayi olurdu ve dogrulugunu kimse denetleyemezdi.
        baglam = baglam_olustur(self.oturum, donem, yalniz_aktif=False)
        kayitlar = atama_kayitlarina_cevir(atamalar)
        kurallar = kurallari_yukle(self.kural.aktif_kurallari_getir())
        h10 = next((k for k in kurallar if k.kimlik == "H10"), None)
        esik = float(h10.parametreler.get("fazla_calisma_esigi", 45)) if h10 else 45.0
        kota = float(h10.parametreler.get("yillik_fazla_kotasi", 270)) if h10 else 270.0
        fazla_calisma = h10_fazla_calisma_saatleri(kayitlar, baglam, esik)
        kalan = {
            p: max(kota - baglam.devir_fazla_calisma_saat(p) - fazla_calisma.get(p, 0.0), 0.0)
            for p in baglam.personel
        }

        aciklar = list(self.kapsama.surume_gore_getir(surum_id))

        # KARSILANMAYAN KISI-SAAT VE CEZA KALEMLERI ANALIZ SERVISINDEN.
        # Ikisi de bir sure burada hesaplaniyordu; SDD 6.3.4 ekran, dosya ve
        # servisin ayni sayiyi vermesini sart kosuyor ve ayni hesabin iki
        # yerde durmasi bu projede birkac kez bedelini odetti. Kalemler
        # (kimlik, ad, ham, agirlik) dortlusune burada yalnizca DONUSTURULUR.
        kalemler = [(k.kimlik, k.ad, k.ham_deger, k.agirlik) for k in analiz.ceza_kalemleri]

        return _Baglam(
            surum=surum,
            donem_baslangic=donem.baslangic_tarihi,
            donem_bitis=donem.bitis_tarihi,
            gunler=gunler,
            atamalar=atamalar,
            aciklar=aciklar,
            fazlalar=list(self.fazla.surume_gore_getir(surum_id)),
            analiz=analiz,
            personel_adi={p.personel_id: p.ad_soyad for p in self.personel.tumunu_getir()},
            personel_sicil={p.personel_id: p.sicil_no for p in self.personel.tumunu_getir()},
            nokta_adi={n.nokta_id: n.ad for n in self.nokta.tumunu_getir()},
            fazla_calisma=fazla_calisma,
            kalan_kota=kalan,
            karsilanmayan_kisi_saat=analiz.karsilanmayan_kisi_saat,
            ceza_kalemleri=kalemler,
        )

    # --- Cizelge calisma kitabi (FR-8.5) ---------------------------------

    def cizelge_calisma_kitabi(self, surum_id: int) -> Workbook | None:
        b = self._baglami_kur(surum_id)
        if b is None:
            return None
        kitap = Workbook()
        self._sayfa_cizelge(kitap.active, b)
        self._sayfa_ozet(kitap.create_sheet("Özet"), b)
        self._sayfa_ham_cizelge(kitap.create_sheet("Ham veri"), b)
        return kitap

    def _basligi_yaz(
        self, sayfa: Worksheet, b: _Baglam, baslik: str, *, olcu_satiri: bool, uretim: bool = True
    ) -> int:
        """Ortak baslik blogu; ilk bos satirin numarasini dondurur.

        `olcu_satiri` yalnizca CIZELGE sayfasinda acilir: kapsama ve acik
        ozeti oraya aittir. Ozet sayfasinda tekrar edilmesi ayni sayiyi iki
        yerde tutmak olurdu ve okuyucuya hicbir sey eklemezdi.
        """
        sayfa["A1"] = baslik
        sayfa["A1"].font = _SAYFA_BASLIGI
        sayfa.row_dimensions[1].height = 25.5
        parcalar = [
            f"Dönem {_tarih_tr(b.donem_baslangic)} – {_tarih_tr(b.donem_bitis)}",
            f"Sürüm {b.surum.surum_no}",
        ]
        if uretim:
            parcalar.append(f"Üretim {_tarih_tr(date.today())}")
        sayfa["A2"] = " · ".join(parcalar)
        sayfa["A2"].font = _ALT_BASLIK
        if not olcu_satiri:
            sayfa.row_dimensions[2].height = 15
            return 4
        sayfa.row_dimensions[2].height = 13.5
        sayfa["A3"] = (
            f"Kapsama {_yuzde_tr(b.analiz.kapsama_orani)} · "
            f"Karşılanmayan {b.karsilanmayan_kisi_saat} kişi-saat "
            f"({len(b.aciklar)} aralık) · "
            f"Talepten fazla {b.analiz.toplam_fazla_kadro}"
        )
        sayfa["A3"].font = _OLCU_YAZI
        sayfa.row_dimensions[3].height = 15
        return 5

    def _sayfa_cizelge(self, sayfa: Worksheet, b: _Baglam) -> None:
        sayfa.title = "Çizelge"
        satir = self._basligi_yaz(sayfa, b, "Vardiya Çizelgesi", olcu_satiri=True)

        # Aciklarin gunleri: gun basligi isaretlenir. Acik SAAT duzeyinde
        # tutulur ama sutun basligi gun duzeyindedir; ayrinti kendi
        # sayfasinda durur.
        acik_gunleri = {a.baslangic_zamani.date() for a in b.aciklar}

        _baslik_satiri(sayfa, satir, ["Personel", *(g.strftime("%d.%m") for g in b.gunler)])
        # ACIK OLAN GUN BASLIGI TURUNCU KALIR. Ortak baslik bandi yesil ama
        # bu isaret bilgi tasiyor ve alttaki aciklama satiri ondan soz ediyor;
        # gorsel butunluk ugruna kaldirilsaydi aciklama yalan soylerdi.
        for sutun, gun in enumerate(b.gunler, start=2):
            if gun in acik_gunleri:
                sayfa.cell(satir, sutun).fill = _ACIK_DOLGU
                sayfa.cell(satir, sutun).font = Font(bold=True, size=10, color=_KOYU_METIN)

        # Blogun SAYILDIGI gun (SRS TD-1) baslangic damgasindan turetilir.
        indeks: dict[tuple[int, date], object] = {
            (a.personel_id, a.baslangic_zamani.date()): a for a in b.atamalar
        }
        personeller = sorted(
            {a.personel_id for a in b.atamalar},
            key=lambda p: b.personel_adi.get(p, ""),
        )
        for i, personel_id in enumerate(personeller, start=1):
            r = satir + i
            _bantla(sayfa, r, len(b.gunler) + 1, i % 2 == 0)
            sayfa.row_dimensions[r].height = 30
            ad = sayfa.cell(r, 1, b.personel_adi.get(personel_id, str(personel_id)))
            ad.font = _KOYU_VERI
            ad.alignment = _SOLA
            ad.border = _KENARLIK
            for sutun, gun in enumerate(b.gunler, start=2):
                hucre = sayfa.cell(r, sutun)
                hucre.border = _KENARLIK
                hucre.alignment = _ORTALI
                atama = indeks.get((personel_id, gun))
                if atama is None:
                    continue
                # METIN HER ZAMAN VAR: dolgu basilmasa da hucre okunur.
                hucre.value = (
                    f"{aralik_metni(atama.baslangic_zamani.time(), atama.bitis_zamani.time())}\n"
                    f"{_nokta_kisaltmasi(b.nokta_adi.get(atama.nokta_id, ''))}"
                )
                hucre.fill, hucre.font = _blok_dolgusu(atama.baslangic_zamani)

        aciklama = satir + len(personeller) + 1
        sayfa.cell(
            aciklama,
            1,
            "Hücre dolgusu çalışmanın gün içindeki konumunu gösterir: koyu = gece "
            "(20.00–06.00), ara ton = kısmen gece, açık = gündüz. Renk tek başına "
            "bilgi taşımaz; saat aralığı hücrede metin olarak da yazılıdır. "
            "Turuncu gün başlığı o günde kapsama açığı bulunduğunu belirtir.",
        ).font = _NOT_YAZI
        # Ikinci not RENGIN NE ANLATTIGINI degil, renk OLMADAN da okunabildigini
        # soyler. Cizelge sahada siyah-beyaz basiliyor; dolgu kaybolunca
        # hucrede yazan saat araligi bilginin TAMAMINI tasimaya devam eder.
        sayfa.cell(
            aciklama + 2,
            1,
            "Renk, çalışmanın günün hangi saatinde başladığını gösterir: koyu gece, "
            "açık gündüz. Saat aralığı hücrede yazılıdır; çıktı renksiz basıldığında "
            "bilgi kaybolmaz.",
        ).font = _NOT_YAZI
        sayfa.column_dimensions["A"].width = 30
        for sutun in range(2, len(b.gunler) + 2):
            sayfa.column_dimensions[get_column_letter(sutun)].width = 16
        sayfa.freeze_panes = sayfa.cell(satir + 1, 2)

    def _sayfa_ozet(self, sayfa: Worksheet, b: _Baglam) -> None:
        satir = self._basligi_yaz(sayfa, b, "Personel Özeti", olcu_satiri=False, uretim=False)
        basliklar = [
            "Sicil",
            "Personel",
            "Toplam saat",
            "Gece saati",
            "Hafta sonu saati",
            "Fazla çalışma",
            "Kalan yıllık kota",
        ]
        _baslik_satiri(sayfa, satir, basliklar)

        # UCU DE ANALIZ SERVISINDEN. Ayni sayilar ekranda da bunlardan
        # okunuyor; burada yeniden toplanmasi iki yuzey arasinda sessiz bir
        # ayrisma acardi.
        gece = {k.personel_id: k.sayi for k in b.analiz.kisi_basina_gece}
        hafta_sonu = {k.personel_id: k.sayi for k in b.analiz.kisi_basina_hafta_sonu}
        for i, denge in enumerate(b.analiz.saat_dagilimi, start=1):
            r = satir + i
            _bantla(sayfa, r, len(basliklar), i % 2 == 0)
            sayfa.cell(r, 1, b.personel_sicil.get(denge.personel_id, ""))
            sayfa.cell(r, 2, denge.ad_soyad)
            sayfa.cell(r, 3, round(denge.toplam_saat, 1))
            sayfa.cell(r, 4, round(gece.get(denge.personel_id, 0.0), 1))
            sayfa.cell(r, 5, round(hafta_sonu.get(denge.personel_id, 0.0), 1))
            sayfa.cell(r, 6, round(b.fazla_calisma.get(denge.personel_id, 0.0), 1))
            sayfa.cell(r, 7, round(b.kalan_kota.get(denge.personel_id, 0.0), 1))

        # TOPLAM satiri CANLI FORMUL. Sabit bir sayi yazsaydik, okuyucu bir
        # satiri silip suzdugunde toplam sessizce yanlis kalirdi. Kalan kota
        # (7. sutun) toplanmaz: kisi basina tavan kalintisidir, toplaminin
        # anlami yok.
        if b.analiz.saat_dagilimi:
            son = satir + len(b.analiz.saat_dagilimi) + 1
            sayfa.cell(son, 2, "TOPLAM").font = _KOYU_VERI
            for sutun in range(3, 7):
                harf = get_column_letter(sutun)
                hucre = sayfa.cell(son, sutun, f"=SUM({harf}{satir + 1}:{harf}{son - 1})")
                hucre.font = _KOYU_VERI
            sayfa.cell(son, 1).font = _KOYU_VERI
        sayfa.freeze_panes = sayfa.cell(satir + 1, 1)
        _genislikleri_ayarla(sayfa, [12, 28, 14, 13, 17, 15, 18])

    def _sayfa_ham_cizelge(self, sayfa: Worksheet, b: _Baglam) -> None:
        """CSV ciktisiyla AYNI icerik (SRS 7.2): blok basina bir satir."""
        basliklar = [
            "tarih",
            "sicil",
            "ad",
            "baslangic",
            "bitis",
            "gorev_noktasi",
            "gece_saat",
            "hafta_sonu_mu",
            "sure_saat",
        ]
        _baslik_satiri(sayfa, 1, basliklar)
        for i, a in enumerate(
            sorted(b.atamalar, key=lambda x: (x.baslangic_zamani, x.personel_id)), start=2
        ):
            gun = a.baslangic_zamani.date()
            sure = round((a.bitis_zamani - a.baslangic_zamani).total_seconds() / 3600)
            sayfa.cell(i, 1, gun.isoformat())
            sayfa.cell(i, 2, b.personel_sicil.get(a.personel_id, ""))
            sayfa.cell(i, 3, b.personel_adi.get(a.personel_id, ""))
            # TAM ISO DAMGASI: gece yarisini asan blogun bitisi ertesi gune
            # duser ve bunu yalnizca damga soyleyebilir (SRS 7.2).
            sayfa.cell(i, 4, a.baslangic_zamani.isoformat())
            sayfa.cell(i, 5, a.bitis_zamani.isoformat())
            sayfa.cell(i, 6, b.nokta_adi.get(a.nokta_id, ""))
            sayfa.cell(i, 7, gece_saat_sayisi(a.baslangic_zamani.time(), a.bitis_zamani.time()))
            sayfa.cell(i, 8, "evet" if gun.weekday() >= 5 else "hayir")
            sayfa.cell(i, 9, sure)
            # HAM VERI KUCUK PUNTOYLA: burasi okunmaz, suzulur ve kopyalanir.
            for sutun in range(1, len(basliklar) + 1):
                sayfa.cell(i, sutun).font = _HAM_YAZI
            sayfa.row_dimensions[i].height = 15
        sayfa.freeze_panes = "A2"
        _genislikleri_ayarla(sayfa, [12, 11, 26, 24, 26, 17, 11, 15, 11])

    # --- Analiz calisma kitabi (FR-8.9) ----------------------------------

    def analiz_calisma_kitabi(self, surum_id: int) -> Workbook | None:
        b = self._baglami_kur(surum_id)
        if b is None:
            return None
        kitap = Workbook()
        self._sayfa_analiz_ozet(kitap.active, b)
        self._sayfa_adalet(kitap.create_sheet("Adalet"), b)
        self._sayfa_aciklar(kitap.create_sheet("Kapsama açıkları"), b)
        self._sayfa_ham_analiz(kitap.create_sheet("Ham veri"), b)
        return kitap

    def _sayfa_analiz_ozet(self, sayfa: Worksheet, b: _Baglam) -> None:
        sayfa.title = "Özet"
        satir = self._basligi_yaz(sayfa, b, "Çizelge Analizi", olcu_satiri=False)

        # Kapsama SAYI olarak yazilir, "%96,9" metni olarak degil: okuyucu
        # dosyada uzerine hesap yapabilsin. Bicimlendirme hucrenin isi.
        # IKI SAYI BUYUK PUNTOYLA: kapsama ve karsilanmayan. Rapora bakan
        # kisinin ilk aradigi ikisi bunlar; geri kalani tabloda.
        sayfa.cell(satir, 1, "Kapsama").font = _ETIKET_YAZI
        oran = sayfa.cell(satir, 2, round(b.analiz.kapsama_orani or 0.0, 3))
        oran.number_format = "0.0%"
        oran.font = _BUYUK_SAYI
        sayfa.row_dimensions[satir].height = 24.45
        sayfa.cell(satir + 1, 1, "Karşılanmayan").font = _ETIKET_YAZI
        sayfa.cell(satir + 1, 2, b.karsilanmayan_kisi_saat).font = _BUYUK_VURGU
        sayfa.row_dimensions[satir + 1].height = 24.45
        sayfa.cell(satir + 2, 1, "Açık aralık sayısı").font = _ETIKET_YAZI
        sayfa.cell(satir + 2, 2, len(b.aciklar)).font = _KOYU_VERI
        sayfa.cell(
            satir + 3,
            1,
            "Karşılanmayan kişi-saat ile açık aralık sayısı farklı ölçülerdir: "
            "ardışık saatler tek kayıtta birleştirilir.",
        ).font = _NOT_YAZI

        # HAM DEGER ve AGIRLIK AYRI SUTUNLARDA, carpim ise CANLI FORMUL.
        # Ikisini tek sutunda toplamak okuyucuya "36 mi 360000 mi" sorusunu
        # birakirdi; birimleri de farkli (kisi-saat / saat / gun karsisinda
        # amac fonksiyonu puani). Formul birakmak agirligi degistirip
        # sonucu dosyada gormeyi de mumkun kilar.
        bas = satir + 5
        _baslik_satiri(sayfa, bas, ["Hedef", "Açıklama", "Ham değer", "Ağırlık", "Ağırlıklı ceza"])
        for i, (kimlik, ad, ham, agirlik) in enumerate(b.ceza_kalemleri, start=1):
            r = bas + i
            _bantla(sayfa, r, 5, i % 2 == 0)
            sayfa.cell(r, 1, kimlik).font = _KOYU_VERI
            sayfa.cell(r, 2, ad)
            sayfa.cell(r, 3, round(ham, 1))
            sayfa.cell(r, 4, agirlik)
            sayfa.cell(r, 5, f"=C{r}*D{r}")
        son = bas + len(b.ceza_kalemleri) + 1
        sayfa.cell(son, 2, "TOPLAM").font = _KOYU_VERI
        if b.ceza_kalemleri:
            sayfa.cell(son, 5, f"=SUM(E{bas + 1}:E{son - 1})").font = _KOYU_VERI
        sayfa.cell(
            son + 2,
            1,
            "Ham değer kuralın kendi biriminde ölçülür (kişi-saat, saat, gün); "
            "ağırlıklı ceza amaç fonksiyonuna giren değerdir. "
            "İkisi aynı sütunda gösterilemez.",
        ).font = _NOT_YAZI
        _genislikleri_ayarla(sayfa, [10, 26, 14, 11, 16])

    def _sayfa_adalet(self, sayfa: Worksheet, b: _Baglam) -> None:
        """Kisi basina gece / hafta sonu / toplam saat, ADIL PAY ve sapma.

        GRAFIKLERIN REFERANS CIZGISI ADIL PAYDIR, havuz ortalamasi degil.
        Ortalamayi gostermek S2'nin acikca reddettigi olcuyu dosyaya tasimak
        olurdu: erisilebilirligi kisitli bir havuz ona gore kalici olarak
        sapmali gorunur. Ayni hata ekranda bir kez yapildi ve Tur 6'da
        duzeltildi; dosyanin onu geri getirmemesi gerekir.
        """
        satir = self._basligi_yaz(sayfa, b, "Adalet Dağılımı", olcu_satiri=False, uretim=False)
        sayfa["A2"] = "Referans çizgi kişiye düşen adil paydır; havuz ortalaması değil (SRS S2)."
        sayfa["A2"].font = _ALT_BASLIK
        basliklar = [
            "Personel",
            "Gece saati",
            "Gece adil pay",
            "Hafta sonu saati",
            "Hafta sonu adil pay",
            "Toplam saat",
            "Toplam adil pay",
            "Sapma",
        ]
        _baslik_satiri(sayfa, satir, basliklar)

        gece = {k.personel_id: k for k in b.analiz.kisi_basina_gece}
        hafta_sonu = {k.personel_id: k for k in b.analiz.kisi_basina_hafta_sonu}
        for sira, denge in enumerate(b.analiz.saat_dagilimi, start=1):
            i = satir + sira
            _bantla(sayfa, i, len(basliklar), sira % 2 == 0)
            g = gece.get(denge.personel_id)
            h = hafta_sonu.get(denge.personel_id)
            sayfa.cell(i, 1, denge.ad_soyad)
            sayfa.cell(i, 2, round(g.sayi, 1) if g else 0)
            sayfa.cell(i, 3, round(g.pay or 0.0, 1) if g else 0)
            sayfa.cell(i, 4, round(h.sayi, 1) if h else 0)
            sayfa.cell(i, 5, round(h.pay or 0.0, 1) if h else 0)
            sayfa.cell(i, 6, round(denge.toplam_saat, 1))
            sayfa.cell(i, 7, round(denge.hedef_saat, 1))
            sayfa.cell(i, 8, round(denge.sapma, 1))

        son_satir = len(b.analiz.saat_dagilimi) + satir
        if son_satir > satir:
            # IKI grafik, ucu degil: TOPLAM SAAT ile GECE adaletin iki ayri
            # sorusudur. Hafta sonu saati tabloda durur ama kendi grafigini
            # hak etmez - ucuncu grafik sayfayi uzatir, okuyucunun karsilastirdigi
            # sey azalmaz. Ikisi de TABLONUN ALTINA, A sutununa yerlesir:
            # yanda duran grafik tablo genisleyince veriyi ortuyordu.
            _adalet_grafigi(
                sayfa, "Toplam saat ve adil pay", 6, 7, satir, son_satir, son_satir + 12
            )
            _adalet_grafigi(sayfa, "Gece saati ve adil pay", 2, 3, satir, son_satir, son_satir + 32)
        sayfa.freeze_panes = sayfa.cell(satir + 1, 2)
        _genislikleri_ayarla(sayfa, [26, 13, 15, 17, 19, 13, 16, 12])

    def _sayfa_aciklar(self, sayfa: Worksheet, b: _Baglam) -> None:
        satir = self._basligi_yaz(sayfa, b, "Kapsama Açıkları", olcu_satiri=False, uretim=False)
        sayfa["A2"] = f"{len(b.aciklar)} aralık · toplam {b.karsilanmayan_kisi_saat} kişi-saat"
        sayfa["A2"].font = _ALT_BASLIK
        basliklar = ["Gün", "Saat aralığı", "Görev noktası", "Eksik kişi", "Kişi-saat"]
        _baslik_satiri(sayfa, satir, basliklar)
        if not b.aciklar:
            # Aciklarin YOKLUGU da bildirilir; bos bir sayfa "acik yok" ile
            # "rapor uretilmedi" arasindaki farki soylemez.
            sayfa.cell(satir + 1, 1, "Bu sürümde kapsama açığı yok.")
        for sira, a in enumerate(b.aciklar, start=1):
            i = satir + sira
            _bantla(sayfa, i, len(basliklar) - 1, sira % 2 == 0)
            # Gun ISO: bu sayfa filtrelenip siralanan bir KAYIT listesidir,
            # okunan bir cizelge degil; ISO siralamasi takvim sirasiyla ortusur.
            sayfa.cell(i, 1, a.baslangic_zamani.date().isoformat())
            # Aralik gun sinirini asabilir (B-23); metin iki damgadan kurulur.
            sayfa.cell(
                i, 2, f"{saat_metni(a.baslangic_zamani.time())}–{saat_metni(a.bitis_zamani.time())}"
            )
            sayfa.cell(i, 3, b.nokta_adi.get(a.nokta_id, ""))
            sayfa.cell(i, 4, a.eksik_sayi)
            # Eksik KISI ile eksik KISI-SAAT ayri sayilardir: 3 kisilik bir
            # saatlik acik ile 1 kisilik uc saatlik acik ayni degildir.
            # KISI-SAAT sutunu vurgulu: sayfanin asil olcusu bu, "eksik kisi"
            # degil. Bantlamanin disinda birakilir ki goz once oraya gitsin.
            sayfa.cell(
                i, 5, _kisi_saat(a.baslangic_zamani, a.bitis_zamani, a.eksik_sayi)
            ).font = _OLCU_YAZI
        if b.aciklar:
            son = satir + len(b.aciklar) + 1
            sayfa.cell(son, 3, "TOPLAM").font = _KOYU_VERI
            sayfa.cell(son, 5, f"=SUM(E{satir + 1}:E{son - 1})").font = _OLCU_YAZI
        sayfa.freeze_panes = sayfa.cell(satir + 1, 1)
        _genislikleri_ayarla(sayfa, [14, 18, 22, 13, 13])

    def _sayfa_ham_analiz(self, sayfa: Worksheet, b: _Baglam) -> None:
        """Yukaridaki tablolarin BICIMLENDIRILMEMIS hali."""
        basliklar = ["olcu", "personel_id", "ad", "deger", "adil_pay"]
        _baslik_satiri(sayfa, 1, basliklar)
        r = 2
        for olcu, kalemler in (
            ("gece_saati", b.analiz.kisi_basina_gece),
            ("hafta_sonu_saati", b.analiz.kisi_basina_hafta_sonu),
        ):
            for k in kalemler:
                sayfa.cell(r, 1, olcu)
                sayfa.cell(r, 2, k.personel_id)
                sayfa.cell(r, 3, k.ad_soyad)
                sayfa.cell(r, 4, k.sayi)
                sayfa.cell(r, 5, k.pay)
                r += 1
        for d in b.analiz.saat_dagilimi:
            sayfa.cell(r, 1, "toplam_saat")
            sayfa.cell(r, 2, d.personel_id)
            sayfa.cell(r, 3, d.ad_soyad)
            sayfa.cell(r, 4, d.toplam_saat)
            sayfa.cell(r, 5, d.hedef_saat)
            r += 1
        # HAM VERI KUCUK PUNTOYLA: burasi okunmaz, suzulur ve kopyalanir.
        for satir in range(2, r):
            for sutun in range(1, len(basliklar) + 1):
                sayfa.cell(satir, sutun).font = _HAM_YAZI
            sayfa.row_dimensions[satir].height = 15
        sayfa.freeze_panes = "A2"
        _genislikleri_ayarla(sayfa, [18, 14, 26, 12, 12])


def _adalet_grafigi(
    sayfa: Worksheet,
    baslik: str,
    deger_sutunu: int,
    pay_sutunu: int,
    baslik_satir: int,
    son_satir: int,
    hedef_satir: int,
) -> None:
    """Olculen deger ve ADIL PAY yan yana iki cubuk.

    Onceki surum adil payi cubuklarin uzerine CIZGI olarak bindiriyordu.
    Cizgi kisiden kisiye zipladigi icin - adil pay kisiye ozeldir (SRS S2),
    sabit bir esik degil - okuyucuyu yaniltan bir "trend" gorunumu veriyordu.
    Yan yana iki cubuk karsilastirmayi kisi bazinda birakir: her personelde
    hangi cubugun uzun oldugu dogrudan gorunur.

    Referans OLCUSU degismedi ve degismemeli: kiyas havuz ortalamasina degil
    kisiye dusen adil paya gore yapilir. Ortalamayi gostermek S2'nin acikca
    reddettigi olcuyu dosyaya tasimak olurdu; erisilebilirligi kisitli bir
    havuz ona gore kalici olarak sapmali gorunur. Ayni hata ekranda bir kez
    yapildi ve Tur 6'da duzeltildi.
    """
    cubuk = BarChart()
    cubuk.type = "col"
    cubuk.grouping = "clustered"
    cubuk.title = baslik
    cubuk.y_axis.title = "saat"
    for sutun in (deger_sutunu, pay_sutunu):
        cubuk.add_data(
            Reference(sayfa, min_col=sutun, min_row=baslik_satir, max_row=son_satir),
            titles_from_data=True,
        )
    cubuk.set_categories(Reference(sayfa, min_col=1, min_row=baslik_satir + 1, max_row=son_satir))
    cubuk.height = 9
    cubuk.width = 20
    # Tablonun ALTINA, A sutununa. Yanda duran grafik personel sayisi
    # buyudukce tablonun uzerine biniyordu.
    sayfa.add_chart(cubuk, f"A{hedef_satir}")


def _genislikleri_ayarla(sayfa: Worksheet, genislikler: list[int]) -> None:
    for i, genislik in enumerate(genislikler, start=1):
        sayfa.column_dimensions[get_column_letter(i)].width = genislik


def dosya_adi(surum: CizelgeSurumu, ek: str) -> str:
    """`cizelge_2026-08-10_surum3.xlsx` — dönem ve sürüm adda durur.

    Ayni klasore inen iki dosya tarayicida "(1)" ile ayrisir ve o ad hangi
    surumun hangi donemi oldugunu soylemez.
    """
    return f"{ek}_surum{surum.surum_no}.xlsx"


__all__ = ["DisaAktarmaServisi", "dosya_adi"]
