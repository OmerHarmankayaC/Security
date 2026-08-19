"""Excel disa aktarma (SDD 5.8; SRS FR-8.5, FR-8.9).

Bu dosyanin kilitledigi iki sozlesme:

  1. **Dosya ile ekran ayni sayiyi verir.** Disa aktarma kendi toplamini
     hesaplamaz; kapsama orani, toplam saat ve adil pay `AnalizServisi`den
     gelir. Ayrisirlarsa hangisinin dogru oldugunu kimse bilemez.
  2. **Gece yarisini asan acik dosyada OKUNUR.** B-23 oncesinde kayit tarih +
     ofsetsiz saat tasidigi icin 22.00-02.00 iki satira boluunuyordu.
"""

import uuid
from datetime import date, datetime, time, timedelta
from decimal import Decimal
from io import BytesIO

import pytest
from openpyxl import load_workbook
from sqlalchemy import select

from app.db import OturumYerel
from app.models.kural import Kural, KuralTipi
from app.models.sonuc import (
    Atama,
    AtamaKaynagi,
    CizelgeSurumu,
    CizelgeSurumuDurumu,
    CozumIsi,
    CozumIsiDurumu,
    Donem,
    KapsamaAcigi,
)
from app.models.tanim import GorevNoktasi, GunTipi, Personel, Talep
from app.services.analiz_servisi import AnalizServisi
from app.services.disa_aktarma_servisi import DisaAktarmaServisi, dosya_adi
from tests.conftest import pg_yoksa_atla, yetkili_istemci

BASLANGIC = date(2026, 4, 6)
BITIS = date(2026, 4, 12)


@pytest.fixture
def senaryo() -> dict:
    """Iki personel, bir nokta, cozulmus bir surum ve GECE YARISINI ASAN acik."""
    pg_yoksa_atla()
    on_ek = uuid.uuid4().hex[:8]
    oturum = OturumYerel()
    try:
        for kimlik, tip, parametreler in (
            ("H10", KuralTipi.ZORUNLU, {"fazla_calisma_esigi": 45, "yillik_fazla_kotasi": 270}),
        ):
            mevcut = oturum.execute(
                select(Kural).where(Kural.kimlik == kimlik)
            ).scalar_one_or_none()
            if mevcut is None:
                oturum.add(Kural(kimlik=kimlik, tip=tip, parametreler=parametreler, agirlik=None))
            else:
                mevcut.parametreler = parametreler
                mevcut.aktif = True

        nokta = GorevNoktasi(ad=f"Güvenlik-{on_ek}")
        oturum.add(nokta)
        oturum.flush()
        personeller = [
            Personel(
                ad_soyad=f"Kişi {i} {on_ek}",
                sicil_no=f"{on_ek}-{i}",
                haftalik_hedef_saat=40,
                aktif_baslangic=BASLANGIC - timedelta(days=400),
            )
            for i in (1, 2)
        ]
        oturum.add_all(personeller)
        oturum.flush()

        # Talep: her gun 08-16 arasi bir kisi. Kapsama oraninin sifirdan
        # farkli olmasi icin gerekli.
        oturum.add(
            Talep(
                nokta_id=nokta.nokta_id,
                gun_tipi=GunTipi.HAFTA_ICI,
                tarih=None,
                baslangic=time(8, 0),
                bitis=time(16, 0),
                gereken_sayi=1,
            )
        )

        donem = Donem(
            baslangic_tarihi=BASLANGIC,
            bitis_tarihi=BITIS,
            tercih_son_tarihi=BASLANGIC - timedelta(days=7),
        )
        oturum.add(donem)
        oturum.flush()
        surum = CizelgeSurumu(
            donem_id=donem.donem_id, surum_no=1, durum=CizelgeSurumuDurumu.COZULDU
        )
        oturum.add(surum)
        oturum.flush()

        # Bir gunduz blogu ve GECE YARISINI ASAN bir blok.
        oturum.add_all(
            [
                Atama(
                    surum_id=surum.surum_id,
                    personel_id=personeller[0].personel_id,
                    baslangic_zamani=datetime.combine(BASLANGIC, time(8, 0)),
                    bitis_zamani=datetime.combine(BASLANGIC, time(16, 0)),
                    nokta_id=nokta.nokta_id,
                    kaynak=AtamaKaynagi.COZUCU,
                ),
                Atama(
                    surum_id=surum.surum_id,
                    personel_id=personeller[1].personel_id,
                    baslangic_zamani=datetime.combine(BASLANGIC, time(22, 0)),
                    bitis_zamani=datetime.combine(BASLANGIC + timedelta(days=1), time(6, 0)),
                    nokta_id=nokta.nokta_id,
                    kaynak=AtamaKaynagi.COZUCU,
                ),
            ]
        )
        # GECE YARISINI ASAN ACIK: 22.00-02.00, tek kayit.
        oturum.add(
            KapsamaAcigi(
                surum_id=surum.surum_id,
                baslangic_zamani=datetime.combine(BASLANGIC, time(22, 0)),
                bitis_zamani=datetime.combine(BASLANGIC + timedelta(days=1), time(2, 0)),
                nokta_id=nokta.nokta_id,
                eksik_sayi=2,
            )
        )
        oturum.commit()
        return {
            "surum_id": surum.surum_id,
            "surum_no": surum.surum_no,
            "nokta_ad": nokta.ad,
            "personel": [p.personel_id for p in personeller],
            "adlar": [p.ad_soyad for p in personeller],
        }
    finally:
        oturum.close()


def _kitap(surum_id: int, tur: str):
    oturum = OturumYerel()
    try:
        servis = DisaAktarmaServisi(oturum)
        return (
            servis.cizelge_calisma_kitabi(surum_id)
            if tur == "cizelge"
            else servis.analiz_calisma_kitabi(surum_id)
        )
    finally:
        oturum.close()


def _hucreler(sayfa) -> list[str]:
    return [str(h.value) for satir in sayfa.iter_rows() for h in satir if h.value is not None]


# --- Cizelge kitabi -------------------------------------------------------


def test_cizelge_kitabi_uc_sayfa_tasir(senaryo: dict) -> None:
    kitap = _kitap(senaryo["surum_id"], "cizelge")
    assert kitap is not None
    assert kitap.sheetnames == ["Çizelge", "Özet", "Ham veri"]


def test_cizelge_hucresi_saat_araligini_metin_olarak_tasir(senaryo: dict) -> None:
    """Renk tek basina bilgi tasimaz: renksiz basilan cikti okunur kalmali."""
    kitap = _kitap(senaryo["surum_id"], "cizelge")
    metinler = _hucreler(kitap["Çizelge"])
    assert any("08.00–16.00" in m for m in metinler)
    # Gece yarisini asan blok da TEK aralik olarak yazilir.
    assert any("22.00–06.00" in m for m in metinler)
    # Dolgunun ne anlama geldigini soyleyen aciklama satiri bulunmali.
    assert any("Hücre dolgusu" in m for m in metinler)


def test_ham_veri_iso_damgasi_tasir(senaryo: dict) -> None:
    """Gece yarisini asan blogun bitisi ERTESI gune duser (SRS 7.2)."""
    kitap = _kitap(senaryo["surum_id"], "cizelge")
    metinler = _hucreler(kitap["Ham veri"])
    # Damga sunucudan TIMESTAMPTZ olarak gelir, yani ofset tasir; testin
    # ilgilendigi sey tarih+saat bolumu ve BITISIN ERTESI GUNE dusmesi.
    assert any(m.startswith(f"{BASLANGIC.isoformat()}T22:00:00") for m in metinler)
    ertesi = (BASLANGIC + timedelta(days=1)).isoformat()
    assert any(m.startswith(f"{ertesi}T06:00:00") for m in metinler)


# --- KABUL: dosya ile ekran ayni sayiyi verir -----------------------------


def test_dosyadaki_sayilar_analiz_ekraniyla_birebir_ayni(senaryo: dict) -> None:
    """SDD 5.8: disa aktarma ikinci bir hesap yapmaz."""
    oturum = OturumYerel()
    try:
        analiz = AnalizServisi(oturum).hesapla(senaryo["surum_id"])
    finally:
        oturum.close()
    assert analiz is not None

    kitap = _kitap(senaryo["surum_id"], "analiz")
    assert kitap is not None
    adalet = kitap["Adalet"]

    # Sayfadaki her satir: ad, gece, gece payi, hs, hs payi, toplam, pay, sapma
    satirlar = {satir[0].value: satir for satir in adalet.iter_rows(min_row=2) if satir[0].value}
    assert satirlar, "Adalet sayfasi bos"

    gece = {k.ad_soyad: k for k in analiz.kisi_basina_gece}
    for denge in analiz.saat_dagilimi:
        satir = satirlar.get(denge.ad_soyad)
        assert satir is not None, f"{denge.ad_soyad} dosyada yok"
        # TOPLAM SAAT ve ADIL PAY ekrandakiyle ayni.
        assert satir[5].value == pytest.approx(round(denge.toplam_saat, 1))
        assert satir[6].value == pytest.approx(round(denge.hedef_saat, 1))
        assert satir[7].value == pytest.approx(round(denge.sapma, 1))
        if denge.ad_soyad in gece:
            g = gece[denge.ad_soyad]
            assert satir[1].value == pytest.approx(round(g.sayi, 1))
            # GECE ADIL PAYI da ekrandan gelir; havuz ORTALAMASI degil.
            assert satir[2].value == pytest.approx(round(g.pay or 0.0, 1))

    # Kapsama orani baslik blogunda, analizle ayni degerde.
    # Analiz ozetinde kapsama SAYI olarak durur (uzerine hesap yapilabilsin),
    # cizelge basliginda ise okunacak METIN olarak. Ikisi de ayni degeri
    # ekrandan alir.
    assert analiz.kapsama_orani is not None
    ozet = kitap["Özet"]
    sayisal = {satir[0].value: satir[1].value for satir in ozet.iter_rows(min_row=1, max_col=2)}
    assert sayisal["Kapsama"] == pytest.approx(analiz.kapsama_orani)

    cizelge_metni = " ".join(_hucreler(_kitap(senaryo["surum_id"], "cizelge")["Çizelge"]))
    beklenen = f"%{analiz.kapsama_orani * 100:.1f}".replace(".", ",")
    assert beklenen in cizelge_metni, "Kapsama yuzdesi Turkce ondalik ayraciyla yazilmali"


def test_cizelge_ozeti_de_ayni_toplam_saati_tasir(senaryo: dict) -> None:
    oturum = OturumYerel()
    try:
        analiz = AnalizServisi(oturum).hesapla(senaryo["surum_id"])
    finally:
        oturum.close()
    assert analiz is not None
    ozet = _kitap(senaryo["surum_id"], "cizelge")["Özet"]
    satirlar = {satir[1].value: satir for satir in ozet.iter_rows(min_row=6) if satir[1].value}
    for denge in analiz.saat_dagilimi:
        satir = satirlar.get(denge.ad_soyad)
        if satir is None:
            continue
        assert satir[2].value == pytest.approx(round(denge.toplam_saat, 1))


# --- KABUL: gece yarisini asan acik dosyada okunur ------------------------


def test_gece_yarisini_asan_acik_tek_satirda_ve_okunur(senaryo: dict) -> None:
    """B-23'un asil kazanci.

    Eski gosterimde 22.00-02.00 gun sinirinda bolunuyor ve dosyada iki ayri
    acik gibi gorunuyordu; hangi gune ait oldugu da belirsiz kaliyordu.
    """
    sayfa = _kitap(senaryo["surum_id"], "analiz")["Kapsama açıkları"]
    # Baslik blogu 1-3, sutun basliklari 4. satirda; veri 5'ten baslar.
    satirlar = [
        [h.value for h in satir]
        for satir in sayfa.iter_rows(min_row=5)
        if satir[0].value and satir[0].value != "TOPLAM"
    ]
    assert len(satirlar) == 1, "Aralik BOLUNMEMELI"
    gun, aralik, nokta, eksik, kisi_saat = satirlar[0][:5]
    # Bu sayfa filtrelenip siralanan bir KAYIT listesidir; gun ISO kalir.
    assert gun == BASLANGIC.isoformat()
    assert aralik == "22.00–02.00"
    assert nokta == senaryo["nokta_ad"]
    assert eksik == 2
    # 2 kisi x 4 saat: eksik KISI ile eksik KISI-SAAT ayri sayilardir.
    assert kisi_saat == 8


def test_acik_yokken_sayfa_bunu_acikca_soyler(senaryo: dict) -> None:
    oturum = OturumYerel()
    try:
        for a in oturum.execute(
            select(KapsamaAcigi).where(KapsamaAcigi.surum_id == senaryo["surum_id"])
        ).scalars():
            oturum.delete(a)
        oturum.commit()
    finally:
        oturum.close()
    sayfa = _kitap(senaryo["surum_id"], "analiz")["Kapsama açıkları"]
    assert "Bu sürümde kapsama açığı yok." in _hucreler(sayfa)


# --- Analiz kitabi --------------------------------------------------------


def test_analiz_kitabi_dort_sayfa_ve_grafik_tasir(senaryo: dict) -> None:
    kitap = _kitap(senaryo["surum_id"], "analiz")
    assert kitap.sheetnames == ["Özet", "Adalet", "Kapsama açıkları", "Ham veri"]
    # Adalet iki soruda olculur: TOPLAM saat ve GECE. Her grafikte olculen
    # deger ile ADIL PAY yan yana iki cubuktur; kiyas havuz ortalamasina degil
    # kisiye dusen paya goredir (SRS S2).
    grafikler = kitap["Adalet"]._charts
    assert len(grafikler) == 2
    for grafik in grafikler:
        assert len(grafik.series) == 2


# --- Uc noktalar ----------------------------------------------------------


def test_uc_noktalar_dosyayi_dogrudan_doner(senaryo: dict) -> None:
    with yetkili_istemci() as istemci:
        for yol, ek in (("cizelge", "cizelge"), ("analiz", "analiz")):
            yanit = istemci.get(f"/api/surum/{senaryo['surum_id']}/{yol}.xlsx")
            assert yanit.status_code == 200
            assert yanit.headers["content-type"].startswith("application/vnd.openxmlformats")
            assert f"{ek}_surum{senaryo['surum_no']}.xlsx" in yanit.headers["content-disposition"]
            # Gercekten acilabilir bir calisma kitabi olmali.
            assert load_workbook(BytesIO(yanit.content)).sheetnames


def test_bulunmayan_surumde_404() -> None:
    pg_yoksa_atla()
    with yetkili_istemci() as istemci:
        assert istemci.get("/api/surum/999999/cizelge.xlsx").status_code == 404
        assert istemci.get("/api/surum/999999/analiz.xlsx").status_code == 404


def test_dosya_adi_donem_ve_surumu_tasir(senaryo: dict) -> None:
    oturum = OturumYerel()
    try:
        surum = oturum.get(CizelgeSurumu, senaryo["surum_id"])
        assert dosya_adi(surum, "cizelge") == f"cizelge_surum{surum.surum_no}.xlsx"
    finally:
        oturum.close()


def test_ceza_tablosu_ham_carpi_agirlik_toplam_cezayi_verir(senaryo: dict) -> None:
    """Dosyadaki formul sutunu SUS PAYI DEGIL, gercek iliskiyi gosterir.

    Ham deger ve agirlik ayri sutunlarda durur; carpimlarinin toplami
    `toplam_ceza`ya esit olmalidir. Esit degilse ham degerler cozucunun
    dokumunden degil baska bir yerden geliyor demektir - bir kez tam da bu
    oldu ve dosya ekranla celisen sayilar tasidi.
    """
    dokum = {"S1": 36.0, "S2": 148.0, "S4": 209.0}
    agirliklar = {"S1": 10000.0, "S2": 10.0, "S4": 1.0}
    # Kurallar KURULUP SONUNDA SILINIR. Kural katalogu global; birakilan bir
    # satir sonraki testlerin gordugu kural kumesini degistirir. Bu tuzaga bir
    # kez dusuldu ve basarisiz test kumesi kosudan kosuya degisti.
    eklenen: list[int] = []
    oturum = OturumYerel()
    try:
        mevcut = {
            k.kimlik: k for k in oturum.execute(select(Kural)).scalars() if k.kimlik in agirliklar
        }
        for kimlik, agirlik in agirliklar.items():
            kural = mevcut.get(kimlik)
            if kural is None:
                kural = Kural(kimlik=kimlik, tip=KuralTipi.ESNEK, parametreler={}, aktif=True)
                oturum.add(kural)
                oturum.flush()
                eklenen.append(kural.kural_id)
            kural.agirlik = Decimal(agirlik)
        oturum.add(
            CozumIsi(
                surum_id=senaryo["surum_id"],
                durum=CozumIsiDurumu.TAMAMLANDI,
                baslangic_zamani=datetime(2026, 4, 6, 9, 0),
                bitis_zamani=datetime(2026, 4, 6, 9, 1),
                zaman_limiti_saniye=60,
                en_iyi_ceza=Decimal(sum(dokum[k] * agirliklar[k] for k in dokum)),
                ceza_dokumu=dokum,
                kural_anlik_goruntu={},
            )
        )
        oturum.commit()
        analiz = AnalizServisi(oturum).hesapla(senaryo["surum_id"])
        assert analiz is not None

        sayfa = DisaAktarmaServisi(oturum).analiz_calisma_kitabi(senaryo["surum_id"])["Özet"]
        satirlar = [
            satir
            for satir in sayfa.iter_rows(min_row=10, max_col=5)
            # Agirlik sutunu dolu olan satirlar tablodur; TOPLAM ve kapanis notu degil.
            if satir[3].value is not None
        ]
        assert satirlar, "Ceza tablosu bos"

        toplam = 0.0
        for satir in satirlar:
            kimlik, _ad, ham, agirlik, carpim = (h.value for h in satir)
            assert ham == pytest.approx(analiz.ceza_dokumu[kimlik]), f"{kimlik} dokumden gelmiyor"
            # Carpim SABIT SAYI DEGIL: agirligi degistiren okuyucu sonucu dosyada gorur.
            assert carpim == f"=C{satir[0].row}*D{satir[0].row}"
            toplam += float(ham) * float(agirlik)
        assert toplam == pytest.approx(float(analiz.toplam_ceza or 0.0))
    finally:
        for kural_id in eklenen:
            oturum.delete(oturum.get(Kural, kural_id))
        oturum.commit()
        oturum.close()


def test_cizelge_hucresi_saat_araligi_ve_nokta_kisaltmasi_tasir(senaryo: dict) -> None:
    """Tam nokta adi hucreye sigmiyor; kisaltma ayirt edici ve TURKCE buyutulmus."""
    sayfa = _kitap(senaryo["surum_id"], "cizelge")["Çizelge"]
    kisaltma = senaryo["nokta_ad"][:3].upper()
    dolu = [
        h.value
        for satir in sayfa.iter_rows(min_row=6, min_col=2)
        for h in satir
        if h.value is not None
    ]
    assert dolu, "Cizelge bos"
    for metin in dolu:
        aralik, nokta = metin.split("\n")
        assert "–" in aralik and nokta == kisaltma


def test_ekran_ile_dosya_ayni_kaynaktan_besleniyor(senaryo: dict) -> None:
    """SDD 6.3.4: ekran, dosya ve servis AYNI sayiyi vermek zorunda.

    Karsilanmayan kisi-saat ile ceza kalemleri bir sure DISA AKTARMA
    SERVISINDE hesaplaniyordu; ekran onlari hic gormuyordu. Ikisi
    `AnalizServisi`e tasindi ve bu test iki yuzeyin ayrisamayacagini
    kilitliyor - ayrisirlarsa hangisinin dogru oldugunu kimse bilemez.
    """
    oturum = OturumYerel()
    try:
        analiz = AnalizServisi(oturum).hesapla(senaryo["surum_id"])
    finally:
        oturum.close()
    assert analiz is not None

    ozet = _kitap(senaryo["surum_id"], "analiz")["Özet"]
    sayisal = {satir[0].value: satir[1].value for satir in ozet.iter_rows(min_row=1, max_col=2)}

    assert sayisal["Karşılanmayan"] == analiz.karsilanmayan_kisi_saat
    assert sayisal["Açık aralık sayısı"] == analiz.acik_aralik_sayisi

    # Ceza tablosundaki her satir servisin kalemiyle birebir.
    dosya_kalemleri = {
        satir[0].value: (satir[2].value, satir[3].value)
        for satir in ozet.iter_rows(min_row=10, max_col=5)
        if satir[3].value is not None
    }
    for kalem in analiz.ceza_kalemleri:
        assert kalem.kimlik in dosya_kalemleri, f"{kalem.kimlik} dosyada yok"
        ham, agirlik = dosya_kalemleri[kalem.kimlik]
        assert ham == pytest.approx(round(kalem.ham_deger, 1))
        assert agirlik == pytest.approx(kalem.agirlik)


def test_ozet_sayfasi_devri_ayri_sutunda_tasir(senaryo: dict) -> None:
    """Devir ile donem fazlasi AYRI sutunlarda ve toplam yalniz dogru sutunlari toplar.

    Tek sutunda durduklarinda, kotasi devirden dolmus bir kisi dosyada
    "fazla calisma 0,0 - kalan kota 5,0" olarak okunuyordu; ekranda ayni
    satir duzeltildi, dosya geride kalmamali (SDD 6.3.4).

    Devirin TOPLAM satirina girmemesi ayri bir sarttir: kisi basina tavan
    kalintisi toplanmaz, tıpkı kalan kota gibi.
    """
    oturum = OturumYerel()
    try:
        analiz = AnalizServisi(oturum).hesapla(senaryo["surum_id"])
    finally:
        oturum.close()
    assert analiz is not None

    ozet = _kitap(senaryo["surum_id"], "cizelge")["Özet"]
    basliklar = [h.value for h in next(ozet.iter_rows(min_row=4, max_row=4))]
    assert "Devir (önceki dönemler)" in basliklar
    assert "Fazla çalışma (bu dönem)" in basliklar

    devirler = {k.personel_id: k.devir_saat for k in analiz.kota_durumu}
    fazlalar = {k.personel_id: k.fazla_calisma_saat for k in analiz.kota_durumu}
    satirlar = {satir[1].value: satir for satir in ozet.iter_rows(min_row=6) if satir[1].value}
    for denge in analiz.saat_dagilimi:
        satir = satirlar.get(denge.ad_soyad)
        if satir is None:
            continue
        assert satir[5].value == pytest.approx(round(devirler[denge.personel_id], 1))
        assert satir[6].value == pytest.approx(round(fazlalar[denge.personel_id], 1))

    # TOPLAM satiri: devir (F) ve kalan kota (H) sutunlarinda formul YOK.
    toplam = next(
        (s for s in ozet.iter_rows(min_row=6) if s[1].value == "TOPLAM"),
        None,
    )
    assert toplam is not None
    assert str(toplam[6].value).startswith("=SUM")
    assert toplam[5].value is None
    assert toplam[7].value is None
